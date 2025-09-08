import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import load_workbook
import os
import shutil
from datetime import datetime

class ExcelSheetDeleterGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel シート削除ツール")
        self.root.geometry("600x500")
        self.root.resizable(True, True)
        
        # 初期ファイルパス
        self.file_path = r"C:\Users\yukik\Desktop\excel\test\sales_demo_data.xlsx"
        self.workbook = None
        self.sheet_vars = {}
        
        self.create_widgets()
        self.load_initial_file()
    
    def create_widgets(self):
        # メインフレーム
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # ファイル選択フレーム
        file_frame = ttk.LabelFrame(main_frame, text="ファイル選択", padding="5")
        file_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # ファイルパス表示
        self.file_path_var = tk.StringVar(value=self.file_path)
        file_entry = ttk.Entry(file_frame, textvariable=self.file_path_var, width=60)
        file_entry.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 5))
        
        # ファイル選択ボタン
        browse_btn = ttk.Button(file_frame, text="参照", command=self.browse_file)
        browse_btn.grid(row=0, column=1)
        
        # 読み込みボタン
        load_btn = ttk.Button(file_frame, text="読み込み", command=self.load_file)
        load_btn.grid(row=0, column=2, padx=(5, 0))
        
        file_frame.columnconfigure(0, weight=1)
        
        # ファイル情報フレーム
        info_frame = ttk.LabelFrame(main_frame, text="ファイル情報", padding="5")
        info_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.info_label = ttk.Label(info_frame, text="ファイルを読み込んでください")
        self.info_label.grid(row=0, column=0, sticky=tk.W)
        
        # シート選択フレーム
        sheet_frame = ttk.LabelFrame(main_frame, text="削除するシートを選択", padding="5")
        sheet_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        
        # スクロール可能なフレーム
        canvas = tk.Canvas(sheet_frame, height=200)
        scrollbar = ttk.Scrollbar(sheet_frame, orient="vertical", command=canvas.yview)
        self.scrollable_frame = ttk.Frame(canvas)
        
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        sheet_frame.columnconfigure(0, weight=1)
        sheet_frame.rowconfigure(0, weight=1)
        
        # 操作ボタンフレーム
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=3, column=0, columnspan=2, pady=(0, 10))
        
        # 全選択/全解除ボタン
        select_all_btn = ttk.Button(button_frame, text="全選択", command=self.select_all)
        select_all_btn.grid(row=0, column=0, padx=(0, 5))
        
        deselect_all_btn = ttk.Button(button_frame, text="全解除", command=self.deselect_all)
        deselect_all_btn.grid(row=0, column=1, padx=(0, 20))
        
        # バックアップ設定
        self.backup_var = tk.BooleanVar(value=True)
        backup_check = ttk.Checkbutton(button_frame, text="バックアップを作成", variable=self.backup_var)
        backup_check.grid(row=0, column=2, padx=(0, 20))
        
        # 削除ボタン
        delete_btn = ttk.Button(button_frame, text="選択したシートを削除", command=self.delete_sheets, style="Accent.TButton")
        delete_btn.grid(row=0, column=3)
        
        # ステータスバー
        self.status_var = tk.StringVar(value="準備完了")
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.grid(row=4, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 0))
        
        # グリッドの重み設定
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(2, weight=1)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
    
    def browse_file(self):
        """ファイル選択ダイアログを開く"""
        file_path = filedialog.askopenfilename(
            title="Excelファイルを選択",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
            initialdir=os.path.dirname(self.file_path)
        )
        if file_path:
            self.file_path_var.set(file_path)
            self.file_path = file_path
    
    def load_initial_file(self):
        """初期ファイルを読み込む"""
        if os.path.exists(self.file_path):
            self.load_file()
        else:
            self.status_var.set(f"初期ファイルが見つかりません: {self.file_path}")
    
    def load_file(self):
        """Excelファイルを読み込む"""
        try:
            file_path = self.file_path_var.get()
            if not os.path.exists(file_path):
                messagebox.showerror("エラー", "ファイルが見つかりません")
                return
            
            self.workbook = load_workbook(file_path)
            self.file_path = file_path
            
            # ファイル情報を更新
            sheet_count = len(self.workbook.sheetnames)
            file_size = os.path.getsize(file_path)
            self.info_label.config(text=f"シート数: {sheet_count}, ファイルサイズ: {file_size:,} bytes")
            
            # シート選択チェックボックスを作成
            self.create_sheet_checkboxes()
            
            self.status_var.set(f"ファイルを読み込みました: {os.path.basename(file_path)}")
            
        except Exception as e:
            messagebox.showerror("エラー", f"ファイルの読み込みに失敗しました:\n{e}")
            self.status_var.set("エラー: ファイル読み込み失敗")
    
    def create_sheet_checkboxes(self):
        """シート選択用のチェックボックスを作成"""
        # 既存のウィジェットを削除
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        
        self.sheet_vars = {}
        
        # 各シート用のチェックボックスを作成
        for i, sheet_name in enumerate(self.workbook.sheetnames):
            var = tk.BooleanVar()
            self.sheet_vars[sheet_name] = var
            
            # チェックボックスとシート名を表示
            frame = ttk.Frame(self.scrollable_frame)
            frame.grid(row=i, column=0, sticky=(tk.W, tk.E), padx=5, pady=2)
            
            checkbox = ttk.Checkbutton(frame, variable=var)
            checkbox.grid(row=0, column=0)
            
            label = ttk.Label(frame, text=f"{sheet_name}")
            label.grid(row=0, column=1, sticky=tk.W, padx=(5, 0))
            
            # アクティブなシートは異なる色で表示
            try:
                if self.workbook.active and hasattr(self.workbook.active, 'title') and sheet_name == self.workbook.active.title:
                    label.config(foreground="blue", font=("TkDefaultFont", 9, "bold"))
                    active_label = ttk.Label(frame, text="(アクティブ)", foreground="blue", font=("TkDefaultFont", 8))
                    active_label.grid(row=0, column=2, sticky=tk.W, padx=(5, 0))
            except AttributeError:
                # アクティブシートの情報が取得できない場合は無視
                pass
            
            frame.columnconfigure(1, weight=1)
        
        self.scrollable_frame.columnconfigure(0, weight=1)
    
    def select_all(self):
        """すべてのシートを選択"""
        for var in self.sheet_vars.values():
            var.set(True)
    
    def deselect_all(self):
        """すべてのシートの選択を解除"""
        for var in self.sheet_vars.values():
            var.set(False)
    
    def delete_sheets(self):
        """選択されたシートを削除"""
        if not self.workbook:
            messagebox.showerror("エラー", "ファイルが読み込まれていません")
            return
        
        # 選択されたシートを取得
        selected_sheets = [sheet_name for sheet_name, var in self.sheet_vars.items() if var.get()]
        
        if not selected_sheets:
            messagebox.showwarning("警告", "削除するシートが選択されていません")
            return
        
        # 残るシートが1つ以上あるかチェック
        remaining_sheets = len(self.workbook.sheetnames) - len(selected_sheets)
        if remaining_sheets < 1:
            messagebox.showerror("エラー", "最低1つのシートは必要です。\nすべてのシートを削除することはできません。")
            return
        
        # 確認ダイアログ
        message = f"以下のシートを削除しますか？\n\n"
        message += "\n".join(f"• {sheet}" for sheet in selected_sheets)
        message += f"\n\n残るシート数: {remaining_sheets}個"
        
        if not messagebox.askyesno("確認", message):
            return
        
        try:
            # バックアップを作成
            if self.backup_var.get():
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_path = self.file_path.replace('.xlsx', f'_backup_{timestamp}.xlsx')
                shutil.copy2(self.file_path, backup_path)
                self.status_var.set(f"バックアップ作成: {os.path.basename(backup_path)}")
            
            # シートを削除
            deleted_count = 0
            for sheet_name in selected_sheets:
                try:
                    self.workbook.remove(self.workbook[sheet_name])
                    deleted_count += 1
                except Exception as e:
                    messagebox.showerror("エラー", f"シート '{sheet_name}' の削除に失敗しました:\n{e}")
            
            # ファイルを保存
            self.workbook.save(self.file_path)
            
            # GUIを更新
            self.create_sheet_checkboxes()
            
            # ファイル情報を更新
            sheet_count = len(self.workbook.sheetnames)
            file_size = os.path.getsize(self.file_path)
            self.info_label.config(text=f"シート数: {sheet_count}, ファイルサイズ: {file_size:,} bytes")
            
            messagebox.showinfo("完了", f"{deleted_count}個のシートを削除しました")
            self.status_var.set(f"削除完了: {deleted_count}個のシート")
            
        except Exception as e:
            messagebox.showerror("エラー", f"削除処理中にエラーが発生しました:\n{e}")
            self.status_var.set("エラー: 削除処理失敗")

def main():
    root = tk.Tk()
    app = ExcelSheetDeleterGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()
